# -*- coding: utf-8 -*-
from typing import KeysView
import requests, sys, os, json
import subprocess as sp
import concurrent.futures
import pandas as pd
from openpyxl import load_workbook, worksheet, styles
import packaging.version as pkvers
from tabulate import tabulate
from utils import Utils

## ---------------------------------------------------------------------------------------------- ##

DEBUG = False

NL = '\n'
WORKERS = 10
TIMEOUT = 5
REQUEST_ARGS = {}
VERS_LEVEL = 2
CURRENT = ' (CURRENT)'
MULTI_EXECUTOR_CLASS = concurrent.futures.ThreadPoolExecutor

## ---------------------------------------------------------------------------------------------- ##

class VersionCompare:

    def __init__(self, level=VERS_LEVEL):
        self.level = level

    def get_version(self, version_str):
        version_str = version_str.strip() if version_str else ''
        if not version_str: return pkvers.Version('0')
        parts = version_str.split('.')
        if len(parts) > self.level:
            return pkvers.Version('.'.join(parts[:self.level]))
        return pkvers.Version(version_str)

    def compare_binary(self, pk1, pk2, comp='<'):
        v_1 = self.get_version(pk1)
        v_2 = self.get_version(pk2)
        if comp=='<':
            return v_1 < v_2
        if comp=='>':
            return v_1 > v_2
        if comp=='==':
            return v_1 == v_2
        if comp=='<=':
            return v_1 <= v_2
        if comp=='>=':
            return v_1 >= v_2
        raise Exception(f'Wrong operator: {comp}')

    def compare_binary_reverse(self, pk1, pk2):
        v_1 = self.get_version(pk1)
        v_2 = self.get_version(pk2)
        if v_1 > v_2:
            return '>'
        if v_1 < v_2:
            return '<'
        return '=='

    def is_equal(self, pk1, pk2):
        return self.compare_binary(pk1, pk2, '==')

    def sort_versions(self, versions):
        vv = list(enumerate([self.get_version(v) for v in versions]))
        values = set(map(lambda x: x[1], vv))
        newlist = sorted([[y for y in vv if y[1]==x] for x in values], key=lambda e: e[0][1])
        return [tuple(x[0] for x in e) if len(e) > 1 else e[0][0] for e in newlist]

    def latest_version(self, versions):
        latest = self.sort_versions(versions)[-1]
        return latest if not isinstance(latest, tuple) else None

## ---------------------------------------------------------------------------------------------- ##
class Package:

    prop_names = ['name', 'author', 'summary', 'homepage', 'latest']

    def __init__(self, pk, version=None, package_cache=None, force_update=False, vcomp_or_level=VERS_LEVEL, on_error=None, no_update_cache=False):
        self.vcomp = vcomp_or_level if isinstance(vcomp_or_level, VersionCompare) else VersionCompare(vcomp_or_level)
        self._version = ''
        self.normalized_version = ''
        if not isinstance(pk, Package):
            self._pkname = pk.lower()
            self.version = version
        else:
            self._pkname = pk.name
            self.version = pk.version
        self.on_error = on_error
        self.package_cache = package_cache
        self.force_update = force_update
        self.no_update_cache = no_update_cache
        self.update_properties(pk.asdict(False) if isinstance(pk, Package) else None)

    @property
    def version(self):
        return self._version

    @version.setter
    def version(self, value):
        self._version = value
        self.normalized_version = str(self.vcomp.get_version(value))

    def _properties_set(self):
        return all(p in self.__dict__ for p in Package.prop_names)

    def _get_pkg_info(self):
        try:
            res = requests.get('https://pypi.org/pypi/{}/json'.format(self._pkname),
                            headers={'Accept': 'application/json'}, timeout=TIMEOUT, **REQUEST_ARGS)
            if res.status_code != 200:
                raise Exception(f'HTTP Error {res.status_code}!{NL}{res.text}')
            resjs = json.loads(res.content)
            return resjs['info']

        except Exception as err:
            if self.on_error:
                self.on_error(err)
            else:
                raise

        return dict()

    def update_properties(self, pkinf=None):
        if DEBUG: print(f'>> PACKAGE "{self._pkname}": UPDATING DATA ...')
        pkinf = pkinf or (self.package_cache.get(self._pkname, None) if self.package_cache else None)

        if self.force_update or pkinf is None or not pkinf.get('homepage', '') or not pkinf.get('latest', ''):
            if DEBUG: print(f'       >> PACKAGE "{self._pkname}": NO DATA FOUND IN CACHE OR FORCED UPDATE! GETTING DATA FROM PYPI ...')
            inf = self._get_pkg_info()
            if DEBUG: print(f'       << PACKAGE "{self._pkname}": PYPI DATA FETCHED')
            pkinf = {'name': inf.get('name', '') or pkinf.get('name', self._pkname) if pkinf else self._pkname,
                     'author': inf.get('author', '') or pkinf.get('author', self._pkname) if pkinf else '',
                     'summary': inf.get('summary', '') or pkinf.get('summary', self._pkname) if pkinf else '',
                     'homepage': inf.get('home_page', inf.get('project_url', inf.get('package_url', ''))) or pkinf.get('homepage', self._pkname) if pkinf else '',
                     'latest': inf.get('version', '') or pkinf.get('latest', self._pkname) if pkinf else ''}

        if not pkinf:
            if self.on_error:
                self.on_error(f'Package {self._pkname} not found on PyPI!')
            else:
                raise Exception(f'Package {self._pkname} not found on PyPI!')

        self.__dict__.update(pkinf)

        if not self.no_update_cache and self.package_cache and self.package_cache.get(self._pkname, {}) != pkinf:
            self.package_cache.update({self._pkname: pkinf})

        if DEBUG: print(f'<< PACKAGE "{self._pkname}": DATA UPDATED')

    def asdict(self, name_as_key=True):
        if not self._properties_set():
            self.update_properties()
        inf = {k: v for k, v in self.__dict__.items() if k in Package.prop_names}
        inf['version'] = self._version
        return {self._pkname: inf} if name_as_key else inf

    def is_outdated(self):
        return self.vcomp.compare_binary(self._version, getattr(self, 'latest', ''))

    def install(self, pyexe=None, upgrade=True, force_version=None):
        if DEBUG: print(f'>> PACKAGE "{self._pkname}": INSTALLING ...')
        args = ['install']
        if upgrade:
            args.append('--upgrade')
        if force_version:
            args.append('--force-reinstall')
            args.append(f'{self._pkname}=={force_version}')
        else:
            args.append(self._pkname)
        res = Utils.pip(args, None, pyexe, self.on_error)
        if DEBUG: print(f'<< PACKAGE "{self._pkname}": INSTALLATION COMPLETE')
        return res

    def uninstall(self, pyexe=None):
        return Utils.pip(['uninstall', '--yes'], self._pkname, pyexe, self.on_error)

    def check(self, pyexe=None):
        return Utils.pip(['check'], self._pkname, pyexe, self.on_error)

    def show(self, pyexe=None, showfiles=False, verbose=False):
        args = ['show']
        if showfiles:
            args.append('--files')
        if verbose:
            args.append('--verbose')
        return Utils.pip(args, self._pkname, pyexe, self.on_error)

    def required_by(self, pyexe=None):
        res = self.show(pyexe)
        if not 'Required-by:' in str(res):
            return None
        res = res.split(NL)
        for line in res:
            if line.startswith('Required-by:'):
                line_ = line.split(':')[1].strip()
                return [p.strip() for p in line_.split(',')]
        return None

    def requires(self, pyexe=None):
        res = self.show(pyexe)
        if not 'Requires:' in str(res):
            return None
        res = res.split(NL)
        for line in res:
            if line.startswith('Requires:'):
                line_ = line.split(':')[1].strip()
                return [p.strip() for p in line_.split(',')]
        return None

    def __str__(self):
        return f'Package {self._pkname}{" [" + self.version + "]" if self.version else ""}'

    def __repr__(self):
        return str(self.asdict())

    def __hash__(self):
        return hash((self._pkname, self.version or ''))

    def __eq__(self, other):
        return self._pkname == other._pkname and self.version == other.version

## ---------------------------------------------------------------------------------------------- ##

class Dframe:

    def asdataframe(self):
        return pd.DataFrame()

    def to_xl(self, filepath='pk.xlsx', df=None):
        df = df if not df is None else self.asdataframe()
        if DEBUG: print(f'>> OUTPUTTING TO EXCEL ("{filepath}") ...')
        df.to_excel(filepath, index_label='packages')
        if DEBUG: print(f'<< SAVED TO EXCEL ("{filepath}")')

    def to_csv(self, filepath='pk.csv', df=None, sep=';'):
        df = df if not df is None else self.asdataframe()
        if DEBUG: print(f'>> OUTPUTTING TO CSV ("{filepath}") ...')
        df.to_csv(filepath, sep=sep, index=False)
        if DEBUG: print(f'<< SAVED TO CSV ("{filepath}")')

    def to_html(self, filepath='pk.html', df=None):
        df = df if not df is None else self.asdataframe()
        if DEBUG: print(f'>> OUTPUTTING TO HTML ("{filepath}") ...')
        with open(filepath, 'w', encoding='utf-8') as file_:
            file_.write(df.to_html(na_rep='', index=False, render_links=True))
        if DEBUG: print(f'<< SAVED TO HTML ("{filepath}")')

    def to_json(self, filepath='pk.json', df=None):
        df = df if not df is None else self.asdataframe()
        if DEBUG: print(f'>> OUTPUTTING TO JSON ("{filepath}") ...')
        with open(filepath, 'w', encoding='utf-8') as file_:
            file_.write(df.to_json(orient='index', indent=2))
        if DEBUG: print(f'<< SAVED TO JSON ("{filepath}")')

    def to_pickle(self, filepath='pk.gz', df=None, compression='infer'):
        df = df if not df is None else self.asdataframe()
        if DEBUG: print(f'>> OUTPUTTING TO PICKLE ("{filepath}") ...')
        df.to_pickle(filepath, compression=compression)
        if DEBUG: print(f'<< SAVED TO PICKLE ("{filepath}")')

    def to_clipboard(self, df=None, excel=True, sep=None):
        df = df if not df is None else self.asdataframe()
        df.to_clipboard(excel, sep, index=False, na_rep='')
        if DEBUG: print('<< SAVED TO CLIPBOARD')

    def to_string(self, df=None):
        df = df if not df is None else self.asdataframe()
        return df.to_string(index=False, na_rep='')

    def to_stringx(self, df=None, tablefmt='fancy_grid', maxwidth=200, filepath=None, **kwargs):
        df = df if not df is None else self.asdataframe()
        if maxwidth:
            maxcolw = maxwidth // len(df.columns)
            df = df.transform(lambda x: x.str.wrap(maxcolw))
        kwargs = kwargs or {}
        if tablefmt:
            kwargs['tablefmt'] = tablefmt
        kwargs['headers'] = 'keys'
        kwargs['showindex'] = False
        if not 'stralign' in kwargs:
            kwargs['stralign'] = 'left'
        s = tabulate(df, **kwargs)
        if filepath:
            if DEBUG: print(f'>> OUTPUTTING TO TEXT FILE ("{filepath}") ...')
            with open(filepath, 'w', encoding='utf-8') as file_:
                file_.write(s)
            if DEBUG: print(f'<< SAVED TO TEXT FILE ("{filepath}")')
        return s

## ---------------------------------------------------------------------------------------------- ##

class Packages(Dframe):

    def __init__(self, packages=None, package_cache=None, force_update=False, vcomp_or_level=VERS_LEVEL, on_error=None):
        self.package_cache = package_cache
        full_packages = packages and isinstance(packages[0], Package)
        if full_packages:
            self.packages = packages.copy()
            self._pknames = [pk.name for pk in self.packages]
        else:
            self._pknames = list(pk for pk in packages) or (list(self.package_cache.keys()) if self.package_cache else [])
            self.packages = []
        self.on_error = on_error
        self.force_update = force_update
        self.vcomp = vcomp_or_level if isinstance(vcomp_or_level, VersionCompare) else VersionCompare(vcomp_or_level)
        self._it = None
        if not full_packages:
            self._collect_packages()        

    def get(self, key):
        if isinstance(key, int):
            return self.packages[key]
        for pk in self.packages:
            if pk.name.lower() == key.lower():
                return pk
        return None

    def asdict(self):
        ds = {}
        for pk in self.packages:
            ds.update(pk.asdict())
        return ds

    # overloaded from DFrame
    def asdataframe(self):
        pkdict = self.asdict()
        if pkdict:
            df = pd.DataFrame.from_dict(pkdict, orient='index')
            return df.reindex(sorted(df.index, key=lambda x: x.lower()))
        return None

    # overloaded from DFrame
    def to_xl(self, filepath='pk.xlsx', df=None):
        df = df if not df is None else self.asdataframe()
        ROWS = len(df) + 1
        COLS = len(df.columns) + 1
        try:
            if DEBUG: print(f'OUTPUTTING TO EXCEL ("{filepath}") ...')
            df.to_excel(filepath, index_label='packages')
            wb = load_workbook(filename=filepath)
            ws = wb.active

            # align first column left
            for col in ws.iter_cols(max_col=1, min_row=2, max_row=ROWS):
                for cell in col:
                    cell.alignment = styles.Alignment(horizontal='left')

            # format as table
            tab = worksheet.table.Table(displayName='Table1', ref=f'a1:{Utils.num2az(COLS)}{ROWS}')
            tab.tableStyleInfo = worksheet.table.TableStyleInfo(name='TableStyleMedium8', showFirstColumn=False,
                                                                showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            if 'Table1' in ws.tables:
                del ws.tables['Table1']
            ws.add_table(tab)

            # adjust col widths
            COLW = {'a': 27, 'b': 18, 'c': 35, 'd': 77, 'e': 44, 'f': 16}
            for i in range(7, COLS + 1):
                COLW[Utils.num2az(i)] = 16
            for c in COLW:
                ws.column_dimensions[c].width = COLW[c]

            # save workbook
            wb.save(filename=filepath)
            if DEBUG: print(f'SAVED TO EXCEL ("{filepath}")')

        except Exception as err:
            print(err)

    def get_fullunion(self, other):
        return self._concat_from(other, '+')

    def update_fullunion(self, other):
        self.packages = self.get_fullunion(other)
        self._pknames = [pk.name for pk in self.packages]
        return self

    def get_union(self, other):
        return self._concat_from(other, '|')

    def update_union(self, other):
        self.packages = self.get_union(other)
        self._pknames = [pk.name for pk in self.packages]
        return self

    def get_intersection(self, other):
        return self._concat_from(other, '&')

    def update_intersection(self, other):
        self.packages = self.get_intersection(other)
        self._pknames = [pk.name for pk in self.packages]
        return self

    def get_difference(self, other):
        return self._concat_from(other, '-')

    def update_difference(self, other):
        self.packages = self.get_difference(other)
        self._pknames = [pk.name for pk in self.packages]
        return self

    def get_symmetric_difference(self, other):
        return self._concat_from(other, '^')

    def update_symmetric_difference(self, other):
        self.packages = self.get_symmetric_difference(other)
        self._pknames = [pk.name for pk in self.packages]
        return self

    def list_outdated(self):
        return (pk for pk in self.packages if pk.is_outdated(self.vcomp))

    def list_uptodate(self):
        return (pk for pk in self.packages if not pk.is_outdated(self.vcomp))

    def install(self, packages=None, pyexe=None, upgrade=True, force_version=False, on_install=None):
        packages = packages or self.packages
        if not packages: return
        if not isinstance(packages[0], Package):
            packages_ = []
            self._collect_packages(packages, packages_)
            packages = packages_
        if DEBUG: print(f'>> INSTALLING PACKAGES ({len(packages)}) ...')
        for pk in packages:
            res = pk.install(pyexe, upgrade, pk.version if force_version and pk.version else None)
            if on_install:
                on_install(pk, res)
        if DEBUG: print('<< INSTALLATION COMPLETE')

    def uninstall(self, packages=None, pyexe=None, on_uninstall=None):
        packages = packages or self.packages
        if not packages: return
        if not isinstance(packages[0], Package):
            packages_ = []
            self._collect_packages(packages, packages_)
            packages = packages_
        if DEBUG: print(f'>> UNINSTALLING PACKAGES ({len(packages)}) ...')
        for pk in packages:
            res = pk.uninstall(pyexe)
            if on_uninstall:
                on_uninstall(pk, res)
        if DEBUG: print('<< UNINSTALLATION COMPLETE')

    def check(self, packages=None, pyexe=None):
        packages = packages or self.packages
        if not packages: return
        if not isinstance(packages[0], Package):
            packages_ = []
            self._collect_packages(packages, packages_)
            packages = packages_
        return Utils.pip(['check'] + [pk.name for pk in packages], None, pyexe, self.on_error)

    def _collect_packages(self, pknames=None, packages=None):
        pknames = pknames if pknames else self._pknames
        if not pknames: return
        packages = packages if packages else self.packages
        if not isinstance(packages, list): return

        packages.clear()
        has_versions = Utils.is_iterable(pknames[0])

        def worker(pkname, version):
            pk = Package(pkname, version, self.package_cache, force_update=self.force_update, vcomp_or_level=self.vcomp, on_error=self.on_error)
            packages.append(pk)

        if DEBUG: print(f'>> COLLECTING PACKAGE INFO FOR {len(packages)} PACKAGES ...')
        with MULTI_EXECUTOR_CLASS(max_workers=WORKERS) as executor:
            futures = {executor.submit(worker, pkname, version): pkname for pkname, version in pknames} if has_versions else \
                      {executor.submit(worker, pkname, None): pkname for pkname in pknames}
            for future in concurrent.futures.as_completed(futures):
                pkname = futures[future]
                try:
                    pk = future.result()
                    if DEBUG: print(f'     << COLLECTED PACKAGE {str(pk)}')
                except Exception as err:
                    if self.on_error:
                        self.on_error(f'{pkname}: {str(err)}')
        if DEBUG: print(f'<< COLLECTED PACKAGE INFO FOR {len(packages)} PACKAGES')

    def _get_merged(self, other, op='+'):
        if op=='+':
            return list(set(self.packages + other.packages))

        elif op=='-':
            ps = []
            for pk1 in self.packages:
                for pk2 in other.packages:
                    if pk1.name == pk2.name:
                        if self.vcomp.compare_binary(pk1.version, pk2.version, '>'):
                            ps.append(pk1)
                        break
                else:
                    ps.append(pk1)
            return ps

        elif op=='&':
            ps = []
            for pk1 in self.packages:
                for pk2 in other.packages:
                    if pk1.name == pk2.name and self.vcomp.is_equal(pk1.version, pk2.version):
                        ps.append(pk1)
                        break
            return ps

        elif op=='|':
            ps = []
            other_ = other.packages.copy()
            for pk1 in self.packages:
                for pk2 in other_:
                    if pk1.name == pk2.name:
                        cmp = self.vcomp.compare_binary_reverse(pk1.version, pk2.version)
                        pk = pk1 if cmp != '<' else pk2
                        ps.append(pk)
                        other_.remove(pk2)
                        break
                else:
                    ps.append(pk1)
            ps += other_
            return ps

        elif op=='^':
            union_ = self._get_merged(other, '|')
            inters_ = self._get_merged(other, '&')
            return list(set(union_) - set(inters_))

        else:
            raise Exception(f'Wrong operator: {op}')

    def _concat_from(self, other, op='+'):
        return Packages(self._get_merged(other, op), self.force_update, self.vcomp, self.on_error)

    def __repr__(self):
        return str(self.asdict())

    def __str__(self):
        return NL.join(str(pk) for pk in self.packages) if self.packages else 'No packages'

    def __getitem__(self, key):
        pk = self.get(key)
        if pk is None:
            raise IndexError
        return pk

    def __len__(self):
        return len(self.packages)

    def __iter__(self):
        self._it = iter(self.packages)
        return self._it

    def __next__(self):
        return next(self._it)

    def __or__(self, other):
        return self.get_union(other)

    def __ior__(self, other):
        return self.update_union(other)

    def __add__(self, other):
        return self.get_fullunion(other)

    def __iadd__(self, other):
        return self.update_fullunion(other)

    def __sub__(self, other):
        return self.get_difference(other)

    def __isub__(self, other):
        return self.update_difference(other)

    def __and__(self, other):
        return self.get_intersection(other)

    def __iand__(self, other):
        return self.update_intersection(other)

    def __xor__(self, other):
        return self.get_symmetric_difference(other)

    def __ixor__(self, other):
        return self.update_symmetric_difference(other)

## ---------------------------------------------------------------------------------------------- ##

class Distro(Packages):

    @staticmethod
    def get_pyexe(pyexe):
        return os.path.abspath(pyexe) if pyexe else sys.executable

    def __init__(self, pyexe=None, alias=None, package_cache=None, append_to_current=CURRENT, force_update=False, vcomp_or_level=VERS_LEVEL, on_error=None):
        self.append_to_current = append_to_current
        self.pyexe = Distro.get_pyexe(pyexe)
        self.alias = alias or f'{self._get_env_version()}'
        if self.append_to_current and self.pyexe == sys.executable:
            self.alias += self.append_to_current
        self.on_error = on_error
        super().__init__(self._list_env_packages(), package_cache, force_update, vcomp_or_level, on_error)        
        if not getattr(self, 'packages', None):
            raise Exception(f'Unable to get packages from environment "{self.pyexe}"!')

    def reread(self):
        self._pknames = self._list_env_packages()
        self._collect_packages()

    def install(self, on_install=None):
        if not getattr(self, 'packages', None): return
        return str(self) + NL + super().install(pyexe=self.pyexe, upgrade=True, on_install=on_install)

    def uninstall(self, packages=None, on_uninstall=None):
        if not getattr(self, 'packages', None): return
        return str(self) + NL + super().uninstall(packages=packages, pyexe=self.pyexe, on_uninstall=on_uninstall)

    def check(self):
        return str(self) + NL + Utils.pip(['check'], None, self.pyexe, self.on_error)

    def asdataframe(self):
        return super().asdataframe().rename(columns={'version': self.alias})

    def _list_env_packages(self):
        if DEBUG: print(f'>> LISTING INSTALLED PACKAGES FOR DISTRO {str(self)} ...')
        out = [tuple(s.strip().split('==')) for s in Utils.execute([self.pyexe, '-m', 'pip', 'list', '--format', 'freeze']).split(NL) if s and '==' in s]
        if DEBUG: print(f'<< LISTED INSTALLED PACKAGES FOR DISTRO {str(self)}')
        return out

    def _get_env_version(self):
        try:
            return Utils.execute([self.pyexe, '-V']).split(' ')[-1].strip()
        except:
            return None

    def __hash__(self):
        return hash((self.pyexe,))

    def __eq__(self, other):
        if isinstance(other, str):
            return self.pyexe.lower() == other.lower()
        return self.pyexe.lower() == other.pyexe.lower()

    def __str__(self):
        return f'{self.alias} @ "{self.pyexe}"'

## ---------------------------------------------------------------------------------------------- ##

class Distros(Dframe):

    def __init__(self, pyexes=None, dbdir=None, save_on_exit=True, append_to_current=CURRENT, force_update=False, vcomp_or_level=VERS_LEVEL, on_error=print):
        self.force_update = force_update
        self.vcomp = vcomp_or_level if isinstance(vcomp_or_level, VersionCompare) else VersionCompare(vcomp_or_level)
        self.on_error = on_error
        self.package_cache = {}
        self.old_package_cache = {}
        self.distros = []
        self._it = None
        self.save_on_exit = save_on_exit
        self.append_to_current = append_to_current
        self.dbdir = dbdir or os.path.dirname(os.path.realpath(__file__))
        self.dbfile = os.path.join(self.dbdir, 'pypkg.json')
        self.load_db()

        if pyexes:
            if Utils.is_iterable(pyexes):
                if isinstance(pyexes, dict):
                    pyexes_ = [(k, v) for k, v in pyexes.items()]
                else:
                    pyexes_ = [tuple(p[:2]) if Utils.is_iterable(p) and len(p) > 1 else (p, None) for p in pyexes]
            else:
                pyexes_ = {pyexes: None}
            # print(pyexes_)
            self._list_envs(pyexes_)
        else:
            self.distros = [Distro(package_cache=self.package_cache, append_to_current=self.append_to_current,
                                   force_update=self.force_update, vcomp_or_level=self.vcomp, on_error=self.on_error)]

    def __del__(self):
        if self._has_updated() and self.save_on_exit:
            self.save_db()

    def _has_updated(self):
        return self.package_cache != self.old_package_cache

    def list_distros(self, asdict=True):
        if not self.distros: return None
        return {d.pyexe: d.alias for d in self.distros} if asdict else [(d.pyexe, d.alias) for d in self.distros]

    def get(self, key):
        if not key:
            for d in self.distros:
                if d.pyexe.lower() == sys.executable.lower():
                    return d
            return None

        if isinstance(key, int):
            return self.distros[key]

        for d in self.distros:
            if d.pyexe.lower() == key.lower() or d.alias == key:
                return d

        return None

    def load_db(self, filepath=None):
        if filepath:
            self.dbfile = os.path.abspath(filepath)
            self.dbdir = os.path.dirname(self.dbfile)
        if DEBUG: print(f'LOADING DB FROM "{self.dbfile}" ...')
        self.package_cache.clear()
        self.old_package_cache.clear()
        if os.path.isfile(self.dbfile):
            self.package_cache = json.load(open(self.dbfile, 'r', encoding='utf-8'))
            self.old_package_cache = self.package_cache.copy()
            if DEBUG: print(f'LOADED {len(self.package_cache)} PACKAGE DEFS')
        elif DEBUG:
            print('NO DB FILE FOUND! (WILL CREATE NEW ON EXIT)')

    def save_db(self, filepath=None):
        if filepath:
            if self.dbfile != filepath:
                self.old_package_cache.clear()
            self.dbfile = os.path.abspath(filepath)
            self.dbdir = os.path.dirname(self.dbfile)
        if not self._has_updated(): return
        if DEBUG: print(f'SAVING DB TO "{self.dbfile}" ...')
        if self.package_cache:
            with open(self.dbfile, 'w', encoding='utf-8') as jsfile:
                json.dump(self.package_cache, jsfile, ensure_ascii=False, indent=2)
            if DEBUG: print(f'SAVED {len(self.package_cache)} PACKAGE DEFS')
        elif DEBUG:
            print('NO PACKAGE DEFS, NO DB CREATED!')

    # overloaded from DFrame
    def asdataframe(self):
        l = len(self.distros)
        if not l: return pd.DataFrame()
        df = self.distros[0].asdataframe().set_index(Package.prop_names)
        if l == 1: 
            return df.reset_index().fillna('').sort_values('name', key=lambda col: col.str.lower()).reset_index(drop=True)
        for d in self.distros[1:]:
            df = df.join(d.asdataframe().set_index(Package.prop_names), how='outer')
        return df.reset_index().fillna('').sort_values('name', key=lambda col: col.str.lower()).reset_index(drop=True)

    # overloaded from DFrame
    def to_xl(self, filepath='pk.xlsx', df=None):
        df = df if not df is None else self.asdataframe()
        ROWS = len(df) + 1
        COLS = len(df.columns) + 1
        try:
            if DEBUG: print(f'OUTPUTTING TO EXCEL ("{filepath}") ...')
            df.to_excel(filepath, index_label='packages')
            wb = load_workbook(filename=filepath)
            ws = wb.active

            # align first column left
            for col in ws.iter_cols(max_col=1, min_row=2, max_row=ROWS):
                for cell in col:
                    cell.alignment = styles.Alignment(horizontal='left')

            # format as table
            tab = worksheet.table.Table(displayName='Table1', ref=f'a1:{Utils.num2az(COLS)}{ROWS}')
            tab.tableStyleInfo = worksheet.table.TableStyleInfo(name='TableStyleMedium8', showFirstColumn=False,
                                                                showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            if 'Table1' in ws.tables:
                del ws.tables['Table1']
            ws.add_table(tab)

            # adjust col widths
            COLW = {'a': 27, 'b': 18, 'c': 35, 'd': 77, 'e': 44, 'f': 16}
            for i in range(7, COLS + 1):
                COLW[Utils.num2az(i)] = 16
            for c in COLW:
                ws.column_dimensions[c].width = COLW[c]

            # highlight missing and latest versions
            for row in ws.iter_rows(min_row=2, max_row=ROWS, min_col=7, max_col=COLS):
                for cell in row:
                    if not cell.value:
                        cell.style = 'Accent2'
                cells = list(row)
                lv = self.vcomp.latest_version([c.value or '' for c in cells])
                if not lv is None:
                    cells[lv].style = 'Accent1'

            # save workbook
            wb.save(filename=filepath)
            if DEBUG: print(f'SAVED TO EXCEL ("{filepath}")')

        except Exception as err:
            Utils.trace_exc()
            # print(err)

    def _list_envs(self, pyexes, on_distro=None):
        def worker(pyexe, alias):
            cnt = sum(1 for d in self.distros if d.alias == alias)
            distro = Distro(pyexe, alias if not cnt else f'{alias}_{cnt}', self.package_cache, self.append_to_current, self.force_update, self.vcomp, self.on_error)
            self.distros.append(distro)
            return distro

        if DEBUG: print(f'>> CREATING DIRTROS ({len(pyexes)}) ...')
        with MULTI_EXECUTOR_CLASS(max_workers=WORKERS) as executor:
            futures = {executor.submit(worker, *pyexe): pyexe for pyexe in pyexes}
            for future in concurrent.futures.as_completed(futures):
                pyexe, alias = futures[future]
                try:
                    distro = future.result()
                    if distro:
                        if DEBUG: print(f'   << CREATED DISTRO {str(distro)}')
                        if on_distro: on_distro(distro)

                except Exception as err:
                    if self.on_error:
                        self.on_error(f'Error retrieving env "{alias}" ("{pyexe}"): {str(err)}')
        self.distros = list(set(self.distros))
        if DEBUG: print(f'<< CREATED DIRTROS ({len(self.distros)})')

    def __getitem__(self, key):
        d = self.get(key)
        if d is None:
            raise IndexError
        return d

    def __str__(self):
        return NL.join(str(d) for d in self.distros)

    def __len__(self):
        return len(self.distros)

    def __iter__(self):
        self._it = iter(self.distros)
        return self._it

    def __next__(self):
        return next(self._it)